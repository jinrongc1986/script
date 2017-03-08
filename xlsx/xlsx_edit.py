#!/usr/bin/env python
# -*- coding:utf-8 -*-

# 加载模块
import sys,time,os
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Color,Fill,Font,PatternFill
#星域提供的节点信息
f=open('university.txt')
lista=[]
for line in f.readlines():
	lista.append(line.strip('\n'))
#本地保存的节点信息
wb = load_workbook('CDS设备信息.xlsx')
sheetname='VPE'
ws=wb[sheetname]
#比对并写入excel
for i in range(len(lista)):
#for i in range(1):
	#print lista[i]
	for row in range(2,90):
		col=5
		value=ws.cell(row=row,column=col).value
		if value==lista[i]:
			ws.cell(row=row,column=8).value='success'
			greenFill = PatternFill(start_color='0032CD32',end_color='0032CD32',fill_type='solid')
			ws.cell(row=row,column=8).fill = greenFill
			break
wb.save('123.xlsx')
